import sqlite3
from datetime import datetime
import os


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    Workbook = None

# --- CONFIGURACIÓN ---
DB_NAME = "fechajusta_v2.db"
LISTA_BOCAS = ["Berchia", "Mayor", "Chango Mas", "Vea", "Carrefour"]

# --- FUNCIONES DE BASE DE DATOS ---

def inicializar_sistema():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS productos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            boca TEXT,
            nombre TEXT,
            marca TEXT, 
            categoria TEXT,
            cantidad INTEGER,
            vencimiento TEXT,
            precio REAL,
            estado TEXT
        )
    """)
    conn.commit()
    conn.close()

def guardar_producto(boca, nombre, marca, cat, cant, venc, precio):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO productos (boca, nombre, marca, categoria, cantidad, vencimiento, precio, estado)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'Pendiente')
    """, (boca, nombre, marca, cat, cant, venc, precio))
    conn.commit()
    conn.close()
    print(f"💾 Guardado: {nombre} ({marca})")

def obtener_productos(filtro_texto=None):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    if filtro_texto:
        parametro = f"%{filtro_texto}%"
        sql = """
            SELECT id, boca, nombre, marca, vencimiento, estado, cantidad, precio 
            FROM productos 
            WHERE (nombre LIKE ? OR marca LIKE ?) 
            ORDER BY vencimiento ASC
        """
        cursor.execute(sql, (parametro, parametro))
    else:
        sql = """
            SELECT id, boca, nombre, marca, vencimiento, estado, cantidad, precio 
            FROM productos 
            ORDER BY vencimiento ASC
        """
        cursor.execute(sql)
        
    datos = cursor.fetchall()
    conn.close()
    return datos

def cambiar_estado(id_prod, nuevo_estado):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("UPDATE productos SET estado = ? WHERE id = ?", (nuevo_estado, id_prod))
    conn.commit()
    conn.close()

def eliminar_producto_db(id_prod):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM productos WHERE id = ?", (id_prod,))
    conn.commit()
    conn.close()

def obtener_estadisticas():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM productos")
    total_items = cursor.fetchone()[0]
    cursor.execute("SELECT SUM(precio * cantidad) FROM productos")
    resultado_dinero = cursor.fetchone()[0]
    total_dinero = resultado_dinero if resultado_dinero else 0
    conn.close()
    return total_items, total_dinero

# --- ESTA ES LA FUNCIÓN QUE TE FALTA ---
def exportar_a_excel():
    # 1. Verificamos si la librería existe
    if not Workbook:
        return "❌ Error: Falta instalar openpyxl"

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    # Traemos los datos
    cursor.execute("SELECT boca, marca, nombre, cantidad, precio, vencimiento, estado FROM productos ORDER BY vencimiento ASC")
    datos = cursor.fetchall()
    conn.close()

    # 2. Creamos el Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Vencimientos"

    # Encabezados
    headers = ["Sucursal", "Marca", "Producto", "Cant", "Precio Unit", "TOTAL ($)", "Vencimiento", "Estado"]
    ws.append(headers)
    
    # Estilo Azul para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    total_general = 0
    hoy = datetime.now().date()

    # 3. Llenamos los datos
    for row in datos:
        boca, marca, nombre, cant, precio, venc, estado = row
        subtotal = cant * precio
        total_general += subtotal
        
        # Traducir Estado a Texto
        estado_texto = "OK"
        try:
            fecha_obj = datetime.strptime(venc, "%Y-%m-%d").date()
            dias = (fecha_obj - hoy).days
            if estado == "Corregido": estado_texto = "SOLUCIONADO"
            elif dias <= 10: estado_texto = "CRITICO (Vence ya)"
            elif dias <= 30: estado_texto = "ALERTA (Este mes)"
            else: estado_texto = "EN FECHA"
        except:
            estado_texto = "ERROR FECHA"

        ws.append([boca, marca, nombre, cant, precio, subtotal, venc, estado_texto])

    # 4. Totales y Guardado
    ws.append([])
    ws.append(["", "", "", "", "TOTAL RIESGO:", total_general])
    ws[f"F{ws.max_row}"].font = Font(bold=True, color="FF0000")

    # Nombre del archivo con fecha y hora para no pisar anteriores
    nombre_archivo = f"Reporte_{datetime.now().strftime('%H-%M-%S')}.xlsx"
    
    # Guardamos en la carpeta actual (Escritorio/FechaJusta)
    ruta_local = os.path.join(os.getcwd(), nombre_archivo)
    
   # --- CÓDIGO ESPECIAL PARA ANDROID ---
    # Intenta guardar directamente en la carpeta Descargas del celular
    ruta_android = f"/storage/emulated/0/Download/{nombre_archivo}"
    
    try:
        wb.save(ruta_android)
        return f"✅ Revisa en Descargas: {nombre_archivo}"
    except:
        # Si falla, intentamos en la carpeta local (por si acaso)
        try:
            ruta_local = os.path.join(os.getcwd(), nombre_archivo)
            wb.save(ruta_local)
            return f"⚠️ Guardado en carpeta App (Oculta): {nombre_archivo}"
        except Exception as e:
            return f"❌ Error guardando: {str(e)}" 
